import openpyxl
from datetime import datetime
import collections
from db_pack.model import *
from abc import ABC, abstractmethod
import random


faculty = ['ФИРТ','АВИЭТ']


class BaseDataProcess(ABC):

    def __init__(self, name_group = None):
        self.__file = None
        self.name_group = name_group
        self.output_data = []

    @abstractmethod
    def parse(self):
        pass

class Parser_Excel(BaseDataProcess):
    def __init__(self, full_path: str, **kwargs):
        BaseDataProcess.__init__(self)
        self.__file = openpyxl.load_workbook(full_path)
        self.name_group = full_path[full_path.find('(')+1 : full_path.find(',')]
        self.sheet = self.__file.active
        self.spec = self.name_group.split('-')[0]
        self.course = self.name_group.split('-')[1][0]
        self.number = self.name_group.split('-')[1][1:]
        self.output_data = []
        self.subjects = kwargs['subjects']
        print(self.subjects)
        self.groups = kwargs['groups']
        self.teachers = kwargs['teachers']
        self.auditorys = kwargs['auditorys']
        self.faculty = faculty[random.randint(0, 1)]
        print()
    def parse(self):
        length = self.sheet.max_row
        schedule = []

        #timetable = TimeTable(faculty=Faculty(name='FIRT'))
        for i in range(2, length + 1):
            value = self.sheet.cell(row=i, column=1).value
            timetable = TimeTable()
            lesson = Lesson()

            group = Group(course=self.course, spec = self.spec, number = self.number, id_faculty=1)
            timetable.group=group
            try:
                date = datetime.strptime(value, "%d.%m.%Y")
                lesson.date = date
                lesson.number = 1
            except TypeError:
                for col in range(2,5):
                    value = self.sheet.cell(row=i, column=col).value
                    if col == 2:
                        pass
                    elif col == 3:
                        type_lesson, name_subject = str(value).split(' ',maxsplit=1)
                        print(name_subject)
                        if name_subject in self.subjects:
                            lesson.id_subject = self.subjects.index(name_subject)+1
                            print()
                        else:
                            subject = Subject(name=name_subject)
                            lesson.subject = subject
                    elif col == 4:
                        val = str(value).split("\n")
                        name_teacher = val[0]
                        auditory_val = val[1].split()[1]

                        if name_teacher in self.teachers:
                            timetable.id_teacher = self.teachers.index(name_teacher)+1
                        else:
                            teacher = Teacher()
                            teacher.name = name_teacher
                            timetable.teacher = teacher

                        if auditory_val in self.auditorys:
                            lesson.id_auditory = self.auditorys.index(auditory_val)+1
                        else:
                            auditory = Auditory()
                            auditory.name = auditory_val
                            lesson.auditory = auditory

                        timetable.lesson = lesson


            schedule.append(timetable)
        self.output_data = schedule


    # def parse_name_group(self, full_path):
    #     name_group = full_path[full_path.find('(') + 1: full_path.find(',')]
    #     information_group = []
    #     information_group.append(name_group.split('-')[0])
    #     self.spec = name_group.split('-')[0]
    #     self.course = name_group.split('-')[1][0]
    #     self.number = name_group.split('-')[1][1:]
    #     self.sheet = self.__file.active
    #     self.output_data = []

